from __future__ import annotations

from copy import copy
from pathlib import Path
import argparse
import sys

import openpyxl


COL_D = 4
COL_E = 5
COPY_COLS = [1, 2, 3, 6]
BRIDGE_COLS = [7, 8, 9, 10, 11]
CULVERT_COLS = [17, 18, 19, 20, 21]

# 把单元格内容转成整数，识别像 "975" 这种字符串。
def as_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        if text.isdigit():
            return int(text)
        try:
            num = float(text)
        except ValueError:
            return None
        if num.is_integer():
            return int(num)
    return None

# 将源行的样式（格式）复制到目标行
def copy_row_style(ws, src_row, dst_row, max_col):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

# 将源行的指定列内容复制到目标行
def copy_context(ws, src_row, dst_row, cols, max_col):
    for col in cols:
        if col > max_col:
            continue
        # 获取目标单元格
        dst_cell = ws.cell(row=dst_row, column=col)
        if dst_cell.value in (None, ""):
            dst_cell.value = ws.cell(row=src_row, column=col).value
            print(f"[copy_context] Copied value to (row={dst_row}, col={col}): {dst_cell.value}")

# 在指定行的 D 列和 E 列写入桩号值
def set_stake_values(ws, row, start_value):
    ws.cell(row=row, column=COL_D).value = str(start_value)
    ws.cell(row=row, column=COL_E).value = str(start_value + 1)

# 用于给新插入的行填充默认值
def fill_new_row_defaults(ws, row):
    for col in BRIDGE_COLS + CULVERT_COLS:
        cell = ws.cell(row=row, column=col)
        if cell.value in (None, ""):
            cell.value = 0

# 扫描 D 列，找出有数值的行
def find_numeric_rows(ws):
    rows = []
    for row in range(1, ws.max_row + 1):
        # 取第 row 行、第 4 列的原始数值，然后转成整数，如果转换失败，则返回 None
        val = as_int(ws.cell(row=row, column=COL_D).value)
        # 如果转换成功，则将行号和数值添加到列表中
        if val is not None:
            rows.append((row, val))
    print("数值行列表", rows) #[(3, 975), (4, 982), (5, 1071), (6, 1076), ……]
    return rows

# 核心逻辑，在 已有编号的行”中间把缺失的编号补齐
def fill_missing_rows(ws):
    # 扫描 D 列，找出有数值的（行号，数值）列表
    numeric_rows = find_numeric_rows(ws)
    print("数值行列表", numeric_rows)
    # 如果找到的行数小于 2，则返回 0, 0，表示没有插入新行，也没有填充已有行。
    if len(numeric_rows) < 2:
        return 0, 0

    max_col = ws.max_column
    inserted = 0
    filled_existing = 0
    # 从最后一行开始，向前遍历，找到相邻的行，计算它们之间的缺失编号数量，然后插入新行，填充缺失编号。
    # 从后往前遍历是为了在需要插入行时避免行号被顶开影响尚未处理的区间：
    for idx in range(len(numeric_rows) - 1, 0, -1):
        row_i, val_i = numeric_rows[idx - 1]
        row_j, val_j = numeric_rows[idx]
        gap = val_j - val_i - 1
        print("第", idx, "组 gap", gap) # 从往前计算D列数值差
        # 不用填充，直接跳过
        if gap <= 0:
            continue

        # 找出 row_i 和 row_j 之间，D 列没有数值的行号，
        candidates = []
        # range是左闭右开区间
        for r in range(row_i + 1, row_j):
            if as_int(ws.cell(row=r, column=COL_D).value) is None:
                candidates.append(r) 
        print("第", idx, "组 candidates", candidates) 

        # 取两者最小值，优先填充空行，空行不够时，插入新行
        fill_count = min(gap, len(candidates))
        print("第", idx, "组 fill_count", fill_count)
        for offset in range(fill_count):
            row = candidates[offset]
            value = val_i + offset + 1
            copy_context(ws, row_i, row, COPY_COLS, max_col)
            set_stake_values(ws, row, value)
        filled_existing += fill_count

        # 还需要插入的新行数
        remaining = gap - fill_count
        print(f"需要插入的新行数: {remaining}")
        if remaining > 0:
            # 在 row_j（下一个编号行）之前插入，插入后，新插入的第一行位于 insert_at 位置
            insert_at = row_j
            ws.insert_rows(insert_at, remaining)
            for offset in range(remaining):
                row = insert_at + offset
                copy_row_style(ws, row_i, row, max_col)
                copy_context(ws, row_i, row, COPY_COLS, max_col)
                set_stake_values(ws, row, val_i + fill_count + offset + 1)
                fill_new_row_defaults(ws, row)
            inserted += remaining

    return inserted, filled_existing

# 选择输入文件
def choose_input_file(path_arg):
    if path_arg:
        path = Path(path_arg)
        if not path.exists():
            print(f"Input file not found: {path}")
            return None
        return path

    # 如果没有提供路径参数，查找当前目录下的所有 .xlsx 文件
    candidates = [
        path
        for path in Path.cwd().glob("*.xlsx")
        if not path.name.startswith("~$")
    ]
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        print("No .xlsx files found in the current directory.")
        return None

    print("Multiple .xlsx files found. Please provide a path:")
    for path in candidates:
        print(f"  - {path.name}")
    return None



def main():
    # 设置命令行参数解析器
    parser = argparse.ArgumentParser(
        description="Insert missing stake rows and fill columns D/E."
    )
    
    # ========== 参数类型说明 ==========
    # 1. 位置参数（Positional Argument）- 没有横线
    #    特点：按位置顺序提供，不需要横线标识
    #    使用：python mian.py data.xlsx
    #    属性名：args.input（直接使用参数名）
    parser.add_argument("input", nargs="?", help="Input .xlsx file path")
    
    # 2. 可选参数（Optional Argument）- 有横线
    #    2.1 短选项（单横线 -）：输入快速，适合常用选项
    #        使用：python mian.py -o result.xlsx
    #    2.2 长选项（双横线 --）：清晰明确，自解释性强
    #        使用：python mian.py --output result.xlsx
    #    注意：当同时提供短选项和长选项时，属性名使用长选项的名字（去掉--）
    #    属性名：args.output（不是 args.o，因为优先使用长选项名）
    parser.add_argument("-o", "--output", help="Output .xlsx file path")
    
    # 3. 只有长选项的参数
    #    使用：python mian.py --sheet Sheet1
    #    属性名：args.sheet（去掉--前缀）
    parser.add_argument("--sheet", help="Worksheet name (default: active sheet)")
    
    # 用于解析命令行参数，将用户输入转换为程序可用的对象。
     = parser.parse_args()

    input_path = choose_input_file(args.input)
    if not input_path:
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    inserted, filled_existing = fill_missing_rows(ws)

    output_path = Path(args.output) if args.output else input_path.with_name(
        f"{input_path.stem}_filled{input_path.suffix}"
    )
    wb.save(output_path)

    print(
        f"Done. Inserted rows: {inserted}, filled existing rows: {filled_existing}."
    )
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
